import openpyxl
from openpyxl.styles import PatternFill


def apply_billing(file_path: str):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=False):
        rent = row[1].value
        utilities = row[2].value
        total = rent + utilities
        row[3].value = total

        if total < 10000:
            row[3].fill = PatternFill(start_color="FFC7CE", fill_type="solid")
        wb.save(file_path)
